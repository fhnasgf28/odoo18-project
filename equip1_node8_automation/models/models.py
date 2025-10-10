import base64
import io
import pytesseract
import numpy as np
import cv2
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from odoo import models, fields, api


class OcrDocument(models.Model):
    _name = 'ocr.document'
    _description = 'OCR Document Processor'
    _inherit = ['mail.thread']

    name = fields.Char(string='Document Name', required=True)
    image_file = fields.Binary(string='Upload Image', required=True, attachment=True)
    image_filename = fields.Char(string='Image Filename')
    raw_text = fields.Text(string='Extracted Text', readonly=True)
    excel_file = fields.Binary(string='Generated Excel', readonly=True, attachment=True)
    excel_filename = fields.Char(string='Excel Filename')
    has_grid_lines = fields.Boolean(string='Has Grid Lines', default=False)

    @api.depends('image_file')
    def _compute_raw_text(self):
        for rec in self:
            if rec.image_file:
                try:
                    image_data = base64.b64decode(rec.image_file)
                    image = Image.open(io.BytesIO(image_data))
                    text = pytesseract.image_to_string(image)
                    rec.raw_text = text
                except Exception as e:
                    rec.raw_text = f"Error during OCR: {str(e)}"
            else:
                rec.raw_text = False

    def action_process_document(self):
        self.ensure_one()
        self._compute_raw_text()
        
        if not self.image_file:
            return self._show_notification('error', 'No image file to process.')
            
        try:
            # Decode image from base64
            image_data = base64.b64decode(self.image_file)
            image_np = np.frombuffer(image_data, np.uint8)
            image = cv2.imdecode(image_np, cv2.IMREAD_COLOR)
            
            # Check if image has grid lines
            has_grid = self._detect_grid_lines(image)
            self.has_grid_lines = has_grid
            
            if has_grid:
                # Process as grid/table
                excel_data = self._process_grid_image(image)
            else:
                # Process as regular text
                excel_data = self._process_text_image()
                
            # Update record with Excel file
            self.write({
                'excel_file': excel_data,
                'excel_filename': f"{self.name}_output.xlsx"
            })
            
            return self._show_notification('success', 'Document processed and Excel file generated.')
        except Exception as e:
            return self._show_notification('error', f'Error processing document: {str(e)}')

    def _detect_grid_lines(self, image):
        """Detect if the image has grid lines like an Excel sheet"""
        try:
            # Convert to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Apply threshold to get binary image
            _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
            
            # Detect horizontal lines
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 1))
            horizontal_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
            
            # Detect vertical lines
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 25))
            vertical_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
            
            # Count lines
            h_lines = cv2.HoughLinesP(horizontal_lines, 1, np.pi/180, threshold=100, minLineLength=100, maxLineGap=10)
            v_lines = cv2.HoughLinesP(vertical_lines, 1, np.pi/180, threshold=100, minLineLength=100, maxLineGap=10)
            
            # Check if we have enough lines to consider it a grid
            h_count = 0 if h_lines is None else len(h_lines)
            v_count = 0 if v_lines is None else len(v_lines)
            
            return h_count >= 3 and v_count >= 3
        except Exception:
            return False

    def _process_grid_image(self, image):
        """Process image with grid lines to extract table structure"""
        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # Apply threshold
        _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
        
        # Detect horizontal and vertical lines
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 1))
        horizontal_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
        
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 25))
        vertical_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
        
        # Combine horizontal and vertical lines to get grid
        grid = cv2.add(horizontal_lines, vertical_lines)
        
        # Find contours in the grid
        contours, _ = cv2.findContours(grid, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        
        # Filter contours to get cells (rectangles)
        cells = []
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            if w > 20 and h > 20 and w < image.shape[1] * 0.9 and h < image.shape[0] * 0.9:
                cells.append((x, y, w, h))
        
        # Sort cells by position (top to bottom, left to right)
        cells.sort(key=lambda c: (c[1], c[0]))
        
        # Group cells into rows based on y-coordinate
        rows = []
        current_row = []
        current_y = -1
        
        for cell in cells:
            x, y, w, h = cell
            if current_y == -1 or abs(y - current_y) < 10:
                current_row.append(cell)
                current_y = y
            else:
                if current_row:
                    # Sort cells in row by x-coordinate
                    current_row.sort(key=lambda c: c[0])
                    rows.append(current_row)
                current_row = [cell]
                current_y = y
        
        if current_row:
            current_row.sort(key=lambda c: c[0])
            rows.append(current_row)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Table"
        
        # Define border style for cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Extract text from each cell and add to Excel
        for row_idx, row in enumerate(rows, start=1):
            for col_idx, cell in enumerate(row, start=1):
                x, y, w, h = cell
                # Extract cell image
                cell_img = gray[y:y+h, x:x+w]
                
                # OCR the cell content
                cell_text = pytesseract.image_to_string(
                    Image.fromarray(cell_img), 
                    config='--psm 6'  # Assume a single uniform block of text
                ).strip()
                
                # Add to Excel with border
                ws.cell(row=row_idx, column=col_idx, value=cell_text).border = thin_border
        
        # Save Excel to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_data = base64.b64encode(excel_buffer.getvalue())
        
        return excel_data

    def _process_text_image(self):
        """Process image as regular text (no grid lines)"""
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Text"
        
        # Split text into lines and add to Excel
        if self.raw_text:
            lines = self.raw_text.split('\n')
            for i, line in enumerate(lines, start=1):
                if line.strip():  # Skip empty lines
                    ws['A{}'.format(i)] = line
        
        # Save Excel to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_data = base64.b64encode(excel_buffer.getvalue())
        
        return excel_data

    def _show_notification(self, type_msg, message):
        """Helper to show notification"""
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': type_msg.capitalize(),
                'message': message,
                'sticky': False,
                'type': type_msg,
                'tag': 'reload'
            }
        }
